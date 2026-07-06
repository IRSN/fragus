"""
compare_rag_runs.py
-------------------
Statistically compare per-question scores between two RAG configurations.

Usage:
    python compare_rag_runs.py <run_A> <run_B>

Example:
    python compare_rag_runs.py baseline baseline_brut

File naming convention:
    ./data/results/score_test_<run_name>_per_question.xlsx

Each file must contain: q_id | <metrics...>
Rows are paired by q_id (same question across both runs).
"""

import sys
import numpy as np
import pandas as pd
from scipy import stats
from scipy.stats import wilcoxon
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────────────────────
# Parameters
# ──────────────────────────────────────────────────────────────────────────────

ALPHA = 0.05

EXCLUDED_COLS = {
    "q_id", "test_intent", "user_input",
    "anchor_expected_grounded", "anchor_expected_ungrounded",
    "anchor_additional_grounded", "anchor_additional_ungrounded",
}

METHODOLOGY = """
══════════════════════════════════════════════════════════════════════
METHODOLOGY — Statistical Comparison of Two RAG Configurations
══════════════════════════════════════════════════════════════════════

Context
-------
Both RAG configurations are evaluated on the same fixed set of N
questions. For each question i and each metric, an individual score
is available, enabling paired testing.

Pairing
-------
Scores are aligned by question identifier (q_id). Each pair
(score_A_i, score_B_i) refers to the same question, cancelling out
question-level variability and yielding higher statistical power than
a comparison of independent group means.

Statistical test: Wilcoxon signed-rank test (two-sided)
--------------------------------------------------------
For each metric, paired differences are computed:
    δᵢ = score_B(i) − score_A(i)

The Wilcoxon signed-rank test (scipy.stats.wilcoxon) is applied
uniformly across all metrics. This non-parametric test makes no
assumption about the distribution of scores — a key advantage given
that RAG evaluation metrics are bounded in [0, 1] and typically
exhibit skewed or bimodal distributions with mass concentrated near
0 and 1.

Effect size: r = Z / √N
  Conventions: |r| < 0.2 negligible · 0.2–0.5 moderate · > 0.5 large

Multiple comparison correction: Benjamini-Hochberg (FDR)
---------------------------------------------------------
Testing K metrics simultaneously inflates the risk of spurious
significant results. The Benjamini-Hochberg procedure (1995) is
applied to control the False Discovery Rate (FDR ≤ α), the expected
proportion of false positives among significant findings — standard
practice in NLP and machine learning research.

BH procedure:
  1. Sort raw p-values p_(1) ≤ … ≤ p_(K)
  2. p_adj_(i) = min_{j ≥ i} ( p_(j) × K / j ),  capped at 1
  3. Significant if p_adj < α = 0.05

Software and libraries
-----------------------
Python 3.x — numpy, scipy, pandas, openpyxl
══════════════════════════════════════════════════════════════════════
"""


# ──────────────────────────────────────────────────────────────────────────────
# Data loading
# ──────────────────────────────────────────────────────────────────────────────

def load_run(run_name: str) -> pd.DataFrame:
    path = f"./data/results/score_test_{run_name}_per_question.xlsx"
    df = pd.read_excel(path)
    if "q_id" not in df.columns:
        raise ValueError(f"{path} must contain a 'q_id' column")
    df["q_id"] = df["q_id"].astype(str)
    return df.set_index("q_id")


# ──────────────────────────────────────────────────────────────────────────────
# Benjamini-Hochberg correction
# ──────────────────────────────────────────────────────────────────────────────

def benjamini_hochberg(p_values: np.ndarray) -> np.ndarray:
    """Return BH-adjusted p-values."""
    n = len(p_values)
    p = np.array(p_values, dtype=float)
    order = np.argsort(p)
    p_sorted = p[order]
    p_adj_sorted = np.minimum.accumulate((p_sorted * n / np.arange(1, n + 1))[::-1])[::-1]
    p_adj = np.empty(n)
    p_adj[order] = np.minimum(p_adj_sorted, 1.0)
    return p_adj


# ──────────────────────────────────────────────────────────────────────────────
# Paired Wilcoxon test
# ──────────────────────────────────────────────────────────────────────────────

def paired_wilcoxon(scores_a: np.ndarray, scores_b: np.ndarray):
    """Return (p_raw, effect_size r)."""
    mask = ~(np.isnan(scores_a) | np.isnan(scores_b))
    delta = (scores_b - scores_a)[mask]
    n = len(delta)

    if n < 4:
        return np.nan, np.nan

    try:
        _, p_raw = wilcoxon(delta, alternative="two-sided", zero_method="wilcox")
        z = abs(stats.norm.ppf(float(p_raw) / 2))
        effect = z / np.sqrt(n)
    except ValueError:
        # All deltas are zero: no difference
        return 1.0, 0.0

    return float(p_raw), float(effect)


# ──────────────────────────────────────────────────────────────────────────────
# Main comparison
# ──────────────────────────────────────────────────────────────────────────────

def compare_runs(run_a: str, run_b: str) -> pd.DataFrame:
    df_a = load_run(run_a)
    df_b = load_run(run_b)

    common_ids = sorted(set(df_a.index) & set(df_b.index))
    if not common_ids:
        raise ValueError("No common questions found (no matching q_id).")

    metrics = sorted((set(df_a.columns) & set(df_b.columns)) - EXCLUDED_COLS)
    k = len(metrics)
    print(f"{len(common_ids)} questions · {k} metrics · α = {ALPHA}\n")

    rows = []
    for metric in metrics:
        a_vals = df_a.loc[common_ids, metric].to_numpy(dtype=float)
        b_vals = df_b.loc[common_ids, metric].to_numpy(dtype=float)

        mean_a = np.nanmean(a_vals)
        mean_b = np.nanmean(b_vals)
        p_raw, effect = paired_wilcoxon(a_vals, b_vals)

        rows.append({
            "metric":        metric,
            f"mean_{run_a}": round(mean_a, 6),
            f"mean_{run_b}": round(mean_b, 6),
            "diff (B−A)":    round(mean_b - mean_a, 6),
            "effect_size r": round(effect, 4) if not np.isnan(effect) else np.nan,
            "p_value_raw":   round(p_raw, 6) if not np.isnan(p_raw) else np.nan,
        })

    result = pd.DataFrame(rows)

    # Apply BH correction
    p_adj = benjamini_hochberg(result["p_value_raw"].to_numpy(dtype=float))
    result["p_value_BH"] = np.round(p_adj, 6)
    result["significant"] = result["p_value_BH"] < ALPHA
    result["best_run"] = result.apply(
        lambda r: "=" if not r["significant"]
                  else (f"↑ {run_b}" if r["diff (B−A)"] > 0 else f"↑ {run_a}"),
        axis=1,
    )

    result = result.sort_values("p_value_BH")

    print(f"Significant (BH): {result['significant'].sum()} / {k}\n")
    print(result.to_string(index=False))
    return result


# ──────────────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────────────

def export_excel(result: pd.DataFrame, run_a: str, run_b: str):
    out_path = f"./data/results/comparison_{run_a}_vs_{run_b}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        result.to_excel(writer, index=False, sheet_name="Comparison")
        ws = writer.sheets["Comparison"]

        header_fill = PatternFill("solid", fgColor="2D6A8F")
        sig_fill    = PatternFill("solid", fgColor="D6F0DC")
        nonsig_fill = PatternFill("solid", fgColor="F5F5F5")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        sig_col_idx = result.columns.get_loc("significant") + 1

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.fill = sig_fill if ws.cell(row=row_idx, column=sig_col_idx).value else nonsig_fill

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    print(f"\n✅  Output file: {out_path}")


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python compare_rag_runs.py <run_A> <run_B>")
        sys.exit(1)

    run_a, run_b = sys.argv[1], sys.argv[2]
    print(METHODOLOGY)
    result = compare_runs(run_a, run_b)
    export_excel(result, run_a, run_b)
